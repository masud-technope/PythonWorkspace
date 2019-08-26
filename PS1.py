price = 10
rating = 4.9
name = "Mosh"
is_published = True
# print(price, rating, name, is_published)

# name = input("What is your name? ")
# print("Hi " +  name)
# age=input("What is your age? ")
# print(name+" is "+ age)
#
# birth_year=input("What is your birth year? ")
# my_age=2019 - float(birth_year)
# print("My age is " + str(my_age) )

# email=''' Hi there!
#   This is Masud Rahman
#   from Usask'''
#
# print(email)

# course ="Hello World programmers!"
# print(course[0:-2])

# first="Masud"
# last="Smith"
# message = first+" ["+last+"] is a coder"
# f_message=f'{first} [{last}] is a coder'

#print(f_message)
#
# f_email=f''' Hi {first} {last}
#  It has been a while since we last talked.
#  Lets catch up sometimes
#  -Masud
#  '''
# print(f_email)

# f_email="masud.rahman@usask.ca"
# print(f_email.replace('Masud','Asad'))
# print ('rahman' in f_email)
# print(len(f_email))
#
# print(f_email.title())

# print(10/3)
# print(10//3)
# print(10%3)
# print(10*3)
# print(10**3)

import  math
#
# x=-10.5
# print(round(x))
# print(abs(x))
#
# y=10.6
# print(math.ceil(y))
# print(math.floor(y))
# print(math.factorial(5))

# count = 0
#
# while count < 5:
#     day = input("How is today? ")
#
#     if day == "hot":
#         print("Drink a lot of water")
#     elif day == "cold":
#         print("Wear warm clothes!")
#     else:
#         print("Its a lovely day!")
#     count += 1

# home_price=1000000
# high_income=False
# is_goog_credit=True
# down_payment=0
#
# if is_published or high_income:
#     down_payment=home_price*(10/100)
#     print("You need to pay "+ str(down_payment))
# else:
#     down_payment = home_price * (20 / 100)
#     print("You need to pay " + str(down_payment))

# i=0
# while i<5:
#     print(i)
#     i+=1
# else:
#     print("out of the range!")

# trial_limit = 5
# trial = 0
#
# while trial < trial_limit:
#     command = input("Enter your command ")
#     if command == "start":
#         print("The cat has started.. go ahead!")
#     elif command == "stop":
#         print("The car has stopped!")
#     elif command == "quit":
#         print("Quiting the game!")
#         break
#     else:
#         print("I do not understand the command")
#     trial += 1
# else:
#     print("Sorry, You failed!")

# for loop

# for item in range(2,10):
#     print(item)
#
# for item in ["masud","asad","mamun","sayed"]:
#     print(item)

# items =[5,2,3,1,6]
# for x_count in items:
#     output=''
#     for i in range(x_count):
#         output+='x'
#     print(output)

# list=["masud","asad","mamun","sayed"]
# 2D list

# matrix=[[1,2,3], [4,5,6], [7,8,9]]
# print(matrix[0][2])

items = [5, 3, 6, 9, 10, 5, 6]
# print(items.pop())
# print(items.pop())
# # items.clear()
# print(items)
# print(items.index(3))

# print(items.count(5))
#
# items2=items.copy()
# #
# items.sort()
# print(items)
# #
# items.reverse()
# print(items)
# #
# print(items2)

# remove duplicates

# temp = []
# for item in items:
#     if item not in temp:
#         temp.append(item)
# print(temp)

# dealing with tuples
# numbers= (1,2,3)
# print(numbers[2])

# unpacking from tuples
# x,y,z=numbers
# print(x,y,z)
#
#
# #unpack the list
# brothers=["Masud","Asad","Mamun"]
# x1,y1,z1 = brothers
# print(x1,y1,z1)

# dealing with dictionary

# customer = {"name": "John Smith", "email": "js@gmail.com", "age": 30}
# print(customer)
# customer["dob"]="1986"
# print(customer["name"]+" "+ str(customer["age"]))
# print(customer.get("dob","1982"))

# digits = {"1": "one", "2": "two", "3": "three", "4": "four", "5": "five", "sad": ":("}
# phone = input("Enter your phone# ")
# translated = ""
# for item in phone:
#     translated += digits.get(item, "!") + " "
# print(translated)

import emoji

# message = input(">")
# print(message.split(" "))
# emojis = {
#     ":)": emoji.emojize(":smile:", use_aliases=True),
#     ":(": emoji.emojize(":cry:", use_aliases=True)
# }

# print(emojis.get(":)"))
# print(emoji.emojize("Let me :smile:", use_aliases=True))

# message = input("How are you? ")
# words = message.split(" ")
# output = ""
# for word in words:
#     output += emojis.get(word) + " "
# print(output)

# def convert2Emoji(message, secondarg):
#     words = message.split(" ")
#     output = ""
#     for word in words:
#         output += emojis.get(word) + " "
#     print(output)
#
#
# message = input("How are you? ")
# convert2Emoji(secondarg="hey", message=message)

# convert2Emoji(message, "hey")
# # arguments vs parameters
# # parameters are the placeholders in the method defs
# # arguments are the actual value provided to the method call
# # positional arguments ... are like C/Java arguments
# # keyword arguments... include the formal parameter keys. to improve the readability of code
# convert2Emoji(message=message, secondarg="hey")

# dealing with exception

# try:
#     age = int(input("Enter your age: "))
#     income = 1000
#     risk = income / age
#     print(age)
# except ValueError:
#     print("Invalid age!")
# except ZeroDivisionError:
#     print ("Divide by zero!")

# class Point:
#
#     def __init__(self, x, y):
#         self.x = x
#         self.y = y
#
#
#
#     def move(self):
#         print("This is Move")
#
#     def draw(self):
#         print("Lets draw a point")
#
#
# point1 = Point(4,5)
# point1.move()
# point1.draw()
# # point1.x = 10
# # point1.y = 20
# print(point1.x * point1.y)
#
# point2 = Point(9,8)
# point2.x = 30
# print(point2.x)
# #
# point3 = Point(5, 6)
# print(point3.x)

# class inheritance

# class Person:
#
#     def __init__(self, name):
#         self.name = name
#
#     def talk(self):
#         print(self.name + " talks!")
#
#
# class Doctor(Person):
#
#     def __int__(self, name):
#         super(name)
#
#
# class Engineer(Person):
#     def __int__(self, name):
#         super(name)
#
#
# masud = Doctor("Masud Rahman")
# masud.talk()
# mamun = Doctor("Mamun")
# mamun.talk()

# importing modules in the workspace
# from converter import fahrenheit_to_celcius
# import lib.converter
#
# print(fahrenheit_to_celcius(189))
# print(lib.converter.celcius_to_fahrenheit(39.8))

# import utils
#
# numbers = [5, 9, 1, 0, -8, 6]
# print(utils.get_max(numbers))
# print(max(numbers))

import io

# f = open("data/sample.txt", "r")
# with open("data/sample.txt", "r") as f:
#     while f.readable():
#         print(f.readline())
#     f.close()

# import random
#
# members=["masud","asad","mamun","sayed","mukta","deepa","bizli"]
#
# for i in range(3):
#     # print(random.random())
#     # print(random.randint(10,20))
#     print(random.choice(members))


# import random
#
#
# class Dice:
#
#     def __init__(self):
#         pass
#
#     def roll(self):
#         sides = [1, 2, 3, 4, 5, 6, 7, 8, 9, 0]
#         x, y = (random.choice(sides), random.choice(sides))
#         # print(x, y)
#         return x, y
#
#
# for i in range(5):
#     dice = Dice()
#     print(dice.roll())

# dealing with files and paths

from pathlib import Path

# p = Path("testdir")
# # print(p.iterdir())
# if p.exists() == False:
#     p.mkdir()
#     print("Directory created")
# elif p.exists() == True:
#     p.rmdir()
#     print("Directory removed")

# p = Path(".")
# for file in p.glob("*.*"):
#     print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("./data/sample-excel.xlsx")
sheet = wb["Sheet1"]
print(sheet.max_row, sheet.max_column)

for row in range(2, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, col)
        print(cell.value)
    new_cell = sheet.cell(row, 5)
    new_cell.value = "checked!"
#
# values = Reference(sheet,
#                       min_row=2,
#                       max_row=sheet.max_row,
#                       min_col=4,
#                       max_col=4)
# print(values)
#
# chart = BarChart()
# chart.add_data(values)
# sheet.add_chart(chart, 'f2')

# wb.save("sample-excel.xlsx")



